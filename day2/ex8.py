import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

filename = 'video2old.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active

print(ws.title)  # vgsales

# # zmiana nazwy arkusza
# ws.title = "Video Games Sales Data"
#
# wb.save(filename)
# wb.close()

sheet_names = wb.sheetnames
print(sheet_names)
# ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']

sheet = wb[sheet_names[1]]  # indek 1, arkusz numer 2
wb.active = wb.index(sheet)

ws = wb.active
print(ws.title)  # Total Sales by Genre

# formatowanie komórek
ws = wb['Video Games Sales Data']

ws['A1'].font = Font(bold=True, size=12)

for cell in ws[1:1]:
    cell.font = Font(bold=True, size=12)

wb.save(filename)
wb.close()

# # dodanie arkusza do pliku
# wb.create_sheet('Empty Sheet')
# print(wb.sheetnames)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year', 'Empty Sheet']
#
# wb.save(filename)
# wb.close()

# usunięcie arkusza z pliku
# wb.remove(wb['Empty Sheet'])
# print(wb.sheetnames)

# wb.save(filename)
# wb.close()

# arkusz nie istnieje
wb = openpyxl.load_workbook(filename)
# ws = wb['Empty Sheet']
# KeyError: 'Worksheet Empty Sheet does not exist.'

# kopiowanie arkusza
wb.copy_worksheet(wb['Video Games Sales Data'])

wb.save(filename)
wb.close()