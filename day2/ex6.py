import openpyxl

# https://fiszki-go.cytr.us/

filename = 'video2old.xlsx'

# wb = openpyxl.load_workbook('video2old.xlsx')
wb = openpyxl.load_workbook(filename)
ws = wb['vgsales']

ws['P1'] = "Averages Sales"
ws['P2'] = '=AVERAGE(K2:K16328)'

wb.save(filename)
wb.close()

ws['Q1'] = "Number od populated cells"
ws['Q2'] = '=COUNTA(E2:E16328)'

wb.save(filename)
wb.close()

ws['S1'] = "Total Sports Sales"
ws['S2'] = '=COUNTIF(E2:E16328, "sports")'

wb.save(filename)
wb.close()

print(ws['S2'].value)  # =COUNTIF(E2:E16328, "sports") wypisze formułę

ws['T1'] = "Total sum of sports sales"
ws['T2'] = '=SUMIF(E2:E16328, "Sports", K2:K16328)'

wb.save(filename)
wb.close()

# zaokrąglenie do najbliższego 25
ws['U1'] = "Rounded sum of Sports Sales"
ws['U2'] = '=CEILING(T2, 25)'

wb.save(filename)
wb.close()

# do 1
ws['V1'] = "Rounded sum of Sports Sales"
ws['V2'] = '=CEILING(T2, 1)'

wb.save(filename)
wb.close()

# w dół
ws['X1'] = 'Floor'
ws['X2'] = '=FLOOR(T2, 25)'

wb.save(filename)
wb.close()

# do 1
ws['W1'] = "Rounded"
ws['W2'] = '=ROUND(T2, 0)'

wb.save(filename)
wb.close()
