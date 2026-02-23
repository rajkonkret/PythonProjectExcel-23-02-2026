import openpyxl

filename = 'video2.xlsx'

# wb = openpyxl.load_workbook('video2.xlsx')
wb = openpyxl.load_workbook(filename)
ws = wb['vgsales']

row_position = 1

for i in range(1, ws.max_row):
    row_position += 1  # row = row + 1

    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_series = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_series

wb.save(filename)
wb.close()
