from openpyxl import Workbook, load_workbook

# pip
# pip install openpyxl

wb = Workbook()  # tworzy szablon pliku excel
ws = wb.active  # ustawiamy aktywny arkusz

ws['A1'] = 42
ws['A2'] = 42
ws['A3'] = 42

ws.append([1, 2, 3])  # lista int
ws.append([4, 5, 6])  # lista int

# zapisanie danych do pliku excel
wb.save('sample.xlsx')
wb.close()

# wczytanie pliku excel
workbook = load_workbook('sample.xlsx')
sheet = workbook.active  # ustawiamy aktywny arkusz
print(sheet)  # <Worksheet "Sheet">

# wypisanie zawartości z komórki A1
print(sheet['A1'])  # <Cell 'Sheet'.A1>
print(sheet['A1'].value)  # 42

print(sheet['A2'].value)  # 42

# pętla
# for - pętla iteracyjna
for i in range(5):  # od 0 do 4
    print(i)

for row in sheet.iter_rows(min_row=1, max_row=5):
    for cell in row:  # kolumny
        print(cell.value)

# 42
# None
# None
# 42
# None
# None
# 42
# None
# None
# 1
# 2
# 3
# 4
# 5
# 6
# None - nie wiem, stan nieokreślony
